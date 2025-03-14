import pandas as pd
from io import BytesIO
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment

# Define colors for highlighting
RED_FILL = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid")
GREEN_FILL = PatternFill(start_color="FF00FF00", end_color="FF00FF00", fill_type="solid")

def highlight_differences_excel(data1, data2, error_details):
    """
    Create a highlighted Excel file showing differences
    """
    try:
        # Create a new Excel file
        output = BytesIO()

        # Create a writer
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Process each sheet in data1
            for sheet in data1["sheet_names"]:
                # Get the dataframe for this sheet
                df1 = data1["data"][sheet]

                # Write the dataframe to the Excel file
                df1.to_excel(writer, sheet_name=sheet, index=False)

                # Get the worksheet
                worksheet = writer.sheets[sheet]

                # Highlight missing sheets
                if sheet in error_details["missing_sheets"]:
                    # Highlight the entire sheet
                    for row in worksheet.iter_rows():
                        for cell in row:
                            cell.fill = RED_FILL

                    # Add a note to the first cell
                    first_cell = worksheet.cell(row=1, column=1)
                    first_cell.comment = Comment("This sheet is missing in file 2", "Comparison Ability")

                # Highlight column differences
                if sheet in error_details["column_differences"]:
                    col_diffs = error_details["column_differences"][sheet]

                    # Highlight missing columns
                    for col_name in col_diffs["missing"]:
                        if col_name in df1.columns:
                            col_idx = df1.columns.get_loc(col_name) + 1  # +1 because openpyxl is 1-indexed
                            for row in range(1, worksheet.max_row + 1):
                                cell = worksheet.cell(row=row, column=col_idx)
                                cell.fill = RED_FILL

                    # Add a note about reordered columns
                    if col_diffs["reordered"]:
                        first_cell = worksheet.cell(row=1, column=1)
                        comment_text = "Column order is different between files"
                        if first_cell.comment:
                            comment_text = f"{first_cell.comment.text}\n{comment_text}"
                        first_cell.comment = Comment(comment_text, "Comparison Ability")

                # Highlight row differences
                if sheet in error_details["row_differences"]:
                    row_diffs = error_details["row_differences"][sheet]

                    # Highlight missing rows
                    for key, row_idx in row_diffs["missing_rows"].items():
                        # Convert to 1-indexed for openpyxl
                        row_idx = int(row_idx) + 2  # +2 for header and 1-indexing

                        # Check if the row exists in the worksheet
                        if row_idx <= worksheet.max_row:
                            for col in range(1, worksheet.max_column + 1):
                                cell = worksheet.cell(row=row_idx, column=col)
                                cell.fill = RED_FILL

                # Highlight value differences
                if sheet in error_details["value_differences"]:
                    value_diffs = error_details["value_differences"][sheet]

                    for diff in value_diffs:
                        if "key" in diff:
                            # Find the row with this key
                            key_col = df1.columns[0]  # Assume first column is key
                            key = diff["key"]

                            # Find the row index
                            try:
                                row_idx = df1.index[df1[key_col].astype(str) == key][0] + 2  # +2 for header and 1-indexing

                                # Find the column index
                                col_name = diff["column"]
                                if col_name in df1.columns:
                                    col_idx = df1.columns.get_loc(col_name) + 1  # +1 for 1-indexing

                                    # Highlight the cell
                                    cell = worksheet.cell(row=row_idx, column=col_idx)
                                    cell.fill = YELLOW_FILL

                                    # Add a comment with the difference
                                    comment_text = f"Value in file 1: {diff['value1']}\nValue in file 2: {diff['value2']}"
                                    cell.comment = Comment(comment_text, "Comparison Ability")
                            except (IndexError, KeyError):
                                # Skip if row not found
                                pass
                        else:
                            # Use row index directly
                            row_idx = diff["row"] + 2  # +2 for header and 1-indexing

                            # Find the column index
                            col_name = diff["column"]
                            if col_name in df1.columns:
                                col_idx = df1.columns.get_loc(col_name) + 1  # +1 for 1-indexing

                                # Highlight the cell
                                cell = worksheet.cell(row=row_idx, column=col_idx)
                                cell.fill = YELLOW_FILL

                                # Add a comment with the difference
                                comment_text = f"Value in file 1: {diff['value1']}\nValue in file 2: {diff['value2']}"
                                cell.comment = Comment(comment_text, "Comparison Ability")

            # Add a summary sheet
            summary_data = []

            # Add missing sheets
            for sheet in error_details["missing_sheets"]:
                summary_data.append(["Sheet", sheet, "Missing in file 2"])

            # Add extra sheets
            for sheet in error_details["extra_sheets"]:
                summary_data.append(["Sheet", sheet, "Extra in file 2"])

            # Add column differences
            for sheet, col_diffs in error_details["column_differences"].items():
                for col in col_diffs["missing"]:
                    summary_data.append(["Column", f"{sheet}.{col}", "Missing in file 2"])

                for col in col_diffs["extra"]:
                    summary_data.append(["Column", f"{sheet}.{col}", "Extra in file 2"])

                if col_diffs["reordered"]:
                    summary_data.append(["Column Order", sheet, "Different between files"])

            # Add row differences
            for sheet, row_diffs in error_details["row_differences"].items():
                if row_diffs["count_diff"]:
                    summary_data.append(["Row Count", sheet, f"{row_diffs['count_diff'][0]} in file 1, {row_diffs['count_diff'][1]} in file 2"])

                for key in row_diffs["missing_rows"]:
                    summary_data.append(["Row", f"{sheet}.{key}", "Missing in file 2"])

                for key in row_diffs["extra_rows"]:
                    summary_data.append(["Row", f"{sheet}.{key}", "Extra in file 2"])

            # Add value differences
            for sheet, value_diffs in error_details["value_differences"].items():
                for diff in value_diffs:
                    if "key" in diff:
                        summary_data.append(["Value", f"{sheet}.{diff['key']}.{diff['column']}", f"{diff['value1']} vs {diff['value2']}"])
                    else:
                        summary_data.append(["Value", f"{sheet}.row{diff['row']}.{diff['column']}", f"{diff['value1']} vs {diff['value2']}"])

            # Create a summary dataframe
            if summary_data:
                summary_df = pd.DataFrame(summary_data, columns=["Type", "Location", "Difference"])
                summary_df.to_excel(writer, sheet_name="Summary", index=False)

        # Return the Excel file as bytes
        output.seek(0)
        return output.getvalue()

    except Exception as e:
        print(f"Error highlighting Excel file: {str(e)}")
        return None

def highlight_differences_csv(data1, data2, error_details):
    """
    Create a highlighted Excel file from CSV showing differences
    """
    try:
        # Create a new Excel file
        output = BytesIO()

        # Create a writer
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Get the dataframe
            df1 = data1["data"]

            # Write the dataframe to the Excel file
            df1.to_excel(writer, sheet_name="Data", index=False)

            # Get the worksheet
            worksheet = writer.sheets["Data"]

            # Highlight column differences
            if "data" in error_details["column_differences"]:
                col_diffs = error_details["column_differences"]["data"]

                # Highlight missing columns
                for col_name in col_diffs["missing"]:
                    if col_name in df1.columns:
                        col_idx = df1.columns.get_loc(col_name) + 1  # +1 because openpyxl is 1-indexed
                        for row in range(1, worksheet.max_row + 1):
                            cell = worksheet.cell(row=row, column=col_idx)
                            cell.fill = RED_FILL

                # Add a note about reordered columns
                if col_diffs["reordered"]:
                    first_cell = worksheet.cell(row=1, column=1)
                    comment_text = "Column order is different between files"
                    first_cell.comment = Comment(comment_text, "Comparison Ability")

            # Highlight row differences
            if "data" in error_details["row_differences"]:
                row_diffs = error_details["row_differences"]["data"]

                # Highlight missing rows
                for key, row_idx in row_diffs["missing_rows"].items():
                    # Convert to 1-indexed for openpyxl
                    row_idx = int(row_idx) + 2  # +2 for header and 1-indexing

                    # Check if the row exists in the worksheet
                    if row_idx <= worksheet.max_row:
                        for col in range(1, worksheet.max_column + 1):
                            cell = worksheet.cell(row=row_idx, column=col)
                            cell.fill = RED_FILL

            # Highlight value differences
            if "data" in error_details["value_differences"]:
                value_diffs = error_details["value_differences"]["data"]

                for diff in value_diffs:
                    if "key" in diff:
                        # Find the row with this key
                        key_col = df1.columns[0]  # Assume first column is key
                        key = diff["key"]

                        # Find the row index
                        try:
                            row_idx = df1.index[df1[key_col].astype(str) == key][0] + 2  # +2 for header and 1-indexing

                            # Find the column index
                            col_name = diff["column"]
                            if col_name in df1.columns:
                                col_idx = df1.columns.get_loc(col_name) + 1  # +1 for 1-indexing

                                # Highlight the cell
                                cell = worksheet.cell(row=row_idx, column=col_idx)
                                cell.fill = YELLOW_FILL

                                # Add a comment with the difference
                                comment_text = f"Value in file 1: {diff['value1']}\nValue in file 2: {diff['value2']}"
                                cell.comment = Comment(comment_text, "Comparison Ability")
                        except (IndexError, KeyError):
                            # Skip if row not found
                            pass
                    else:
                        # Use row index directly
                        row_idx = diff["row"] + 2  # +2 for header and 1-indexing

                        # Find the column index
                        col_name = diff["column"]
                        if col_name in df1.columns:
                            col_idx = df1.columns.get_loc(col_name) + 1  # +1 for 1-indexing

                            # Highlight the cell
                            cell = worksheet.cell(row=row_idx, column=col_idx)
                            cell.fill = YELLOW_FILL

                            # Add a comment with the difference
                            comment_text = f"Value in file 1: {diff['value1']}\nValue in file 2: {diff['value2']}"
                            cell.comment = Comment(comment_text, "Comparison Ability")

            # Add a summary sheet
            summary_data = []

            # Add column differences
            if "data" in error_details["column_differences"]:
                col_diffs = error_details["column_differences"]["data"]

                for col in col_diffs["missing"]:
                    summary_data.append(["Column", col, "Missing in file 2"])

                for col in col_diffs["extra"]:
                    summary_data.append(["Column", col, "Extra in file 2"])

                if col_diffs["reordered"]:
                    summary_data.append(["Column Order", "data", "Different between files"])

            # Add row differences
            if "data" in error_details["row_differences"]:
                row_diffs = error_details["row_differences"]["data"]

                if row_diffs["count_diff"]:
                    summary_data.append(["Row Count", "data", f"{row_diffs['count_diff'][0]} in file 1, {row_diffs['count_diff'][1]} in file 2"])

                for key in row_diffs["missing_rows"]:
                    summary_data.append(["Row", f"Key: {key}", "Missing in file 2"])

                for key in row_diffs["extra_rows"]:
                    summary_data.append(["Row", f"Key: {key}", "Extra in file 2"])

            # Add value differences
            if "data" in error_details["value_differences"]:
                value_diffs = error_details["value_differences"]["data"]

                for diff in value_diffs:
                    if "key" in diff:
                        summary_data.append(["Value", f"{diff['key']}.{diff['column']}", f"{diff['value1']} vs {diff['value2']}"])
                    else:
                        summary_data.append(["Value", f"row{diff['row']}.{diff['column']}", f"{diff['value1']} vs {diff['value2']}"])

            # Create a summary dataframe
            if summary_data:
                summary_df = pd.DataFrame(summary_data, columns=["Type", "Location", "Difference"])
                summary_df.to_excel(writer, sheet_name="Summary", index=False)

        # Return the Excel file as bytes
        output.seek(0)
        return output.getvalue()

    except Exception as e:
        print(f"Error highlighting CSV file: {str(e)}")
        return None